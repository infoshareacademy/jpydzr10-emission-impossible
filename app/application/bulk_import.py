"""
Hurtowy import danych emisyjnych z plików Excel (.xlsx).

Obsługuje import rekordów do dowolnej tabeli emisyjnej:
- Walidacja Pydantic każdego wiersza
- Automatyczne nadawanie ID (next_id)
- Raport błędów — które wiersze nie przeszły walidacji
- Obsługa wyłącznie plików Excel (openpyxl)
"""

import os
from decimal import Decimal, InvalidOperation
from typing import Optional
from openpyxl import load_workbook

from emissions.models import EmissionFactor
from emissions.models import (
    StationaryCombustion,
    MobileCombustion,
    FugitiveEmission,
    ProcessEmission,
    EnergyConsumption,
)
from pydantic import ValidationError


TABLE_MODELS = {
    "stationary": StationaryCombustion,
    "mobile": MobileCombustion,
    "fugitive": FugitiveEmission,
    "process": ProcessEmission,
    "energy_consumption": EnergyConsumption,
}

SKIP_FIELDS = {"id"}


def _read_excel_rows(file_path: str, sheet_name: Optional[str] = None) -> list[dict]:
    """Wczytuje wiersze z pliku Excel (.xlsx)."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows_iter)]

    rows = []
    for row_values in rows_iter:
        row = {}
        for h, v in zip(headers, row_values):
            if not h:
                continue
            if v is None:
                row[h] = None
            elif isinstance(v, float):
                row[h] = str(Decimal(str(v)))
            else:
                row[h] = str(v).strip()
        rows.append(row)

    wb.close()
    return rows


def _parse_value(value: str):
    """Parsuje wartość z Excel — puste stringi → None."""
    if value == "" or value is None:
        return None
    return value


def bulk_import(file_path: str, repo_name: str, repo,
                sheet_name: Optional[str] = None) -> dict:
    """Importuje rekordy z pliku Excel do repozytorium."""

    if repo_name not in TABLE_MODELS:
        return {"imported": 0, "errors": [(0, f"Nieznana tabela: {repo_name}")], "skipped": 0}

    model_class = TABLE_MODELS[repo_name]

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in (".xlsx", ".xls"):
        return {
            "imported": 0,
            "errors": [(0, "Obsługiwane są tylko pliki XLSX")],
            "skipped": 0,
        }

    raw_rows = _read_excel_rows(file_path, sheet_name)

    imported = 0
    errors = []
    skipped = 0

    for row_num, raw_row in enumerate(raw_rows, start=2):
        row = {k: _parse_value(v) for k, v in raw_row.items() if k.lower() not in SKIP_FIELDS}

        # --- NORMALIZACJA energy_type ---
        ENERGY_TYPE_MAP = {
            "energia elektryczna z oze": "Energia elektryczna z OZE",
            "energia elektryczna nie oze": "Energia elektryczna nie OZE",
            "ciepło z oze": "Ciepło z OZE",
            "ciepło nie oze": "Ciepło nie OZE",
            "chłód z oze": "Chłód z OZE",
            "chłód nie oze": "Chłód nie OZE",
            "para techniczna z oze": "Para Techniczna z OZE",
            "para techniczna nie oze": "Para Techniczna nie OZE",
        }

        raw_type = (row.get("energy_type") or "").strip().lower()
        row["energy_type"] = ENERGY_TYPE_MAP.get(raw_type, row.get("energy_type"))
        # --------------------------------

        # --- KONWERSJA LICZB ---
        for field in ("amount", "emission_tco2eq", "factor"):
            if field in row and row[field] is not None:
                try:
                    row[field] = Decimal(str(row[field]))
                except (InvalidOperation, ValueError):
                    pass

        row["id"] = repo.next_id()

        # --- KONWERSJA JEDNOSTEK DO MWh ---
        UNIT_CONVERSION_TO_MWH = {
            "MWh": 1,
            "kWh": Decimal("0.001"),
            "GJ": Decimal("0.277778"),
            "MJ": Decimal("0.000277778"),
        }

        unit = row.get("unit") or "MWh"
        amount = row.get("amount") or Decimal("0")

        amount_mwh = Decimal(str(amount)) * UNIT_CONVERSION_TO_MWH.get(unit, Decimal("1"))
        row["amount_mwh"] = amount_mwh
        # -----------------------------------

        # --- LICZENIE EMISJI ---
        if repo_name == "energy_consumption":
            year = row.get("year")
            country = row.get("company_country") or row.get("country") or "Polska"
            energy_type = row.get("energy_type")

            factor_obj = EmissionFactor.objects.filter(
                year=year,
                country=country,
                factor_name=energy_type
            ).first()

            if factor_obj:
                try:
                    row["emission_tco2eq"] = amount_mwh * Decimal(str(factor_obj.factor))
                except Exception:
                    row["emission_tco2eq"] = Decimal("0")
            else:
                row["emission_tco2eq"] = Decimal("0")
        # -----------------------------------

        try:
            record = model_class(**row)
        except ValidationError as e:
            for err in e.errors():
                field = " -> ".join(str(loc) for loc in err["loc"])
                errors.append((row_num, f"Pole '{field}': {err['msg']}"))
            skipped += 1
            continue
        except Exception as e:
            errors.append((row_num, f"Nieoczekiwany błąd: {e}"))
            skipped += 1
            continue

        ok, msg = repo.add(record)
        if ok:
            imported += 1
        else:
            errors.append((row_num, f"Błąd zapisu: {msg}"))
            skipped += 1

    return {"imported": imported, "errors": errors, "skipped": skipped}
