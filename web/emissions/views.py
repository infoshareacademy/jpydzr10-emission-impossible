import openpyxl
from calculator.calculation import calculate_record_emissions
from companies.models import Companies
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from django.views.generic import (
    CreateView,
    DeleteView,
    ListView,
    TemplateView,
    UpdateView,
    View,
    FormView,
)

from .forms import (
    EmissionFactorForm,
    EnergyConsumptionForm,
    EnergyConsumptionImportForm,
    EnergyProducedForm,
    EnergyProducedImportForm,
    EnergyPurchasedForm,
    EnergyPurchasedImportForm,
    EnergySoldForm,
    EnergySoldImportForm,
    FugitiveEmissionForm,
    MobileCombustionForm,
    ProcessEmissionForm,
    StationaryCombustionForm,
)
from .models import (
    EmissionFactor,
    EnergyConsumption,
    EnergyProduced,
    EnergyPurchased,
    EnergySold,
    FugitiveEmission,
    MobileCombustion,
    ProcessEmission,
    StationaryCombustion,
)


class Scope2CreateMixin(FormView):
    """
    Wspólna logika dla wszystkich widoków dodawania i aktualizacji z Zakresu 2.
    Automatycznie oblicza emisję przy zapisie.
    """

    template_name = "emissions/energy_form.html"

    def form_valid(self, form):
        instance = form.save(commit=False)
        is_new = instance.pk is None

        # Jeśli edycja — zachowaj stare wartości emisji i wskaźników
        if not is_new:
            try:
                db_instance = self.model.objects.get(pk=instance.pk)
                for field_name in [
                    'calculated_emission_tco2eq',
                    'applied_factor_value',
                    'applied_factor_unit',
                    'applied_converter_value',
                    'applied_converter_unit',
                ]:
                    setattr(instance, field_name, getattr(db_instance, field_name, None))
            except self.model.DoesNotExist:
                pass

        # Track użytkownika
        if is_new:
            instance.created_by = self.request.user

        instance.updated_by = self.request.user

        # Oblicz emisję
        try:
            calculate_record_emissions(instance)
        except ValidationError as e:
            instance.calculated_emission_tco2eq = None
            messages.warning(self.request, f"Zapisano rekord, ale nie wyliczono emisji. Powód: {e}")

        instance.save()

        action_text = "dodano wpis do" if is_new else "zaktualizowano wpis w"
        messages.success(self.request, f"Pomyślnie {action_text}: {self.model._meta.verbose_name}")

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_verbose_name'] = self.model._meta.verbose_name
        return context


class Scope2ListMixin(ListView):
    """
    Uniwersalny widok listy dla wszystkich kategorii Scope 2.
    Obsługuje paginację, filtrowanie, sortowanie i eksport.
    """

    paginate_by = 10
    template_name = "emissions/energy_list.html"
    default_sort = "-year"

    def get_queryset(self):
        qs = self.model.objects.all().select_related('company').order_by(self.default_sort)

        self.current_sort = self.request.GET.get('sort', self.default_sort)
        self.current_year = self.request.GET.get('year', '')
        self.current_company = self.request.GET.get('company', '')

        if self.current_year:
            qs = qs.filter(year=self.current_year)
        if self.current_company:
            qs = qs.filter(company__icontains=self.current_company)

        return qs.order_by(self.current_sort)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'model_verbose_name': self.model._meta.verbose_name,
            'model_verbose_name_plural': self.model._meta.verbose_name_plural,
            'current_sort': self.current_sort,
            'current_year': self.current_year,
            'current_company': self.current_company,
            'add_url_name': f'emissions:{self.model._meta.model_name}-add',
            'edit_url_name': f'emissions:{self.model._meta.model_name}-edit',
            'delete_url_name': f'emissions:{self.model._meta.model_name}-delete',
        })
        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel()
        return super().get(request, *args, **kwargs)

    def export_to_excel(self):
        """Eksportuje do pliku .xlsx"""
        qs = self.get_queryset()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(self.model._meta.verbose_name_plural)[:31]

        fields = [f for f in self.model._meta.fields if f.name not in ['id']]
        ws.append([f.verbose_name.title() for f in fields])

        for obj in qs:
            row = []
            for field in fields:
                value = getattr(obj, field.name)
                if hasattr(obj, f'get_{field.name}_display'):
                    value = getattr(obj, f'get_{field.name}_display')()
                row.append(str(value) if value is not None else '')
            ws.append(row)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'export_{self.model._meta.model_name}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class Scope2DeleteMixin(DeleteView):
    """Wspólna logika dla usuwania rekordów z Zakresu 2."""

    def delete(self, request, *args, **kwargs):
        model_verbose = self.model._meta.verbose_name
        messages.success(
            self.request,
            f"Pomyślnie usunięto wpis z: {model_verbose}"
        )
        return super().delete(request, *args, **kwargs)


# ===== ENERGY CONSUMPTION =====
class EnergyConsumptionListView(Scope2ListMixin):
    model = EnergyConsumption

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Zużycie energii')
        return context


class EnergyConsumptionCreateView(Scope2CreateMixin, CreateView):
    model = EnergyConsumption
    form_class = EnergyConsumptionForm
    success_url = reverse_lazy('energy_consumption_list')


class EnergyConsumptionUpdateView(Scope2CreateMixin, UpdateView):
    model = EnergyConsumption
    form_class = EnergyConsumptionForm
    success_url = reverse_lazy('energy_consumption_list')


class EnergyConsumptionDeleteView(Scope2DeleteMixin, DeleteView):
    model = EnergyConsumption
    success_url = reverse_lazy('energy_consumption_list')


# ===== ENERGY PURCHASED =====
class EnergyPurchasedListView(Scope2ListMixin):
    model = EnergyPurchased

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Zakupiona energia')
        return context


class EnergyPurchasedCreateView(Scope2CreateMixin, CreateView):
    model = EnergyPurchased
    form_class = EnergyPurchasedForm
    success_url = reverse_lazy('energy_purchased_list')


class EnergyPurchasedUpdateView(Scope2CreateMixin, UpdateView):
    model = EnergyPurchased
    form_class = EnergyPurchasedForm
    success_url = reverse_lazy('energy_purchased_list')


class EnergyPurchasedDeleteView(Scope2DeleteMixin, DeleteView):
    model = EnergyPurchased
    success_url = reverse_lazy('energy_purchased_list')


# ===== ENERGY PRODUCED =====
class EnergyProducedListView(Scope2ListMixin):
    model = EnergyProduced

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Wyprodukowana energia')
        return context


class EnergyProducedCreateView(Scope2CreateMixin, CreateView):
    model = EnergyProduced
    form_class = EnergyProducedForm
    success_url = reverse_lazy('energy_produced_list')


class EnergyProducedUpdateView(Scope2CreateMixin, UpdateView):
    model = EnergyProduced
    form_class = EnergyProducedForm
    success_url = reverse_lazy('energy_produced_list')


class EnergyProducedDeleteView(Scope2DeleteMixin, DeleteView):
    model = EnergyProduced
    success_url = reverse_lazy('energy_produced_list')


# ===== ENERGY SOLD =====
class EnergySoldListView(Scope2ListMixin):
    model = EnergySold

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Sprzedana energia')
        return context


class EnergySoldCreateView(Scope2CreateMixin, CreateView):
    model = EnergySold
    form_class = EnergySoldForm
    success_url = reverse_lazy('energy_sold_list')


class EnergySoldUpdateView(Scope2CreateMixin, UpdateView):
    model = EnergySold
    form_class = EnergySoldForm
    success_url = reverse_lazy('energy_sold_list')


class EnergySoldDeleteView(Scope2DeleteMixin, DeleteView):
    model = EnergySold
    success_url = reverse_lazy('energy_sold_list')


class Scope1CreateMixin(LoginRequiredMixin):
    """
    Wspólna logika dla wszystkich widoków dodawania i aktualizacji z Zakresu 1.
    Używa tego samego szablonu formularza i wraca na dashboard.
    """

    template_name = (
        "emissions/scope1_form.html"  # Jeden uniwersalny szablon dla formularzy
    )
    url_name_prefix = None

    def get_url_prefix(self):
        return self.url_name_prefix or self.model._meta.model_name

    def get_success_url(self):
        prefix = self.get_url_prefix()
        return reverse_lazy(
            f"emissions:{prefix}-list",
            kwargs={"company_id": self.kwargs.get("company_id")},
        )

    def form_valid(self, form):
        instance = form.save(commit=False)
        instance.company_id = self.kwargs.get("company_id")

        is_new = instance.pk is None

        if not is_new:
            try:
                db_instance = self.model.objects.get(pk=instance.pk)
                for field_name in (
                    "calculated_emission_tco2eq",
                    "applied_factor_value",
                    "applied_factor_unit",
                    "applied_converter_value",
                    "applied_converter_unit",
                ):
                    setattr(
                        instance, field_name, getattr(db_instance, field_name, None)
                    )
            except self.model.DoesNotExist:
                pass

        if is_new:
            instance.created_by = self.request.user

        instance.updated_by = self.request.user

        try:
            calculate_record_emissions(instance)
        except ValidationError as e:
            instance.calculated_emission_tco2eq = None
            instance.factor_used = None
            messages.warning(
                self.request, f"Zapisano rekord, ale nie wyliczono emisji. Powód: {e}"
            )

        instance.save()

        action_text = "dodano wpis do" if is_new else "zaktualizowano wpis w"
        messages.success(
            self.request, f"Pomyślnie {action_text}: {self.model._meta.verbose_name}"
        )

        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["model_verbose_name"] = self.model._meta.verbose_name
        context["company_id"] = self.kwargs.get("company_id")
        prefix = self.get_url_prefix()
        context["list_url_name"] = f"emissions:{prefix}-list"
        context["company"] = get_object_or_404(
            Companies, pk=self.kwargs.get("company_id")
        )
        return context


class EmissionListMixin(LoginRequiredMixin, ListView):
    """
    Uniwersalny widok listy dla wszystkich kategorii emisji.
    Obsługuje paginację, filtrowanie, sortowanie i eksport do XLSX.
    """

    paginate_by = 15
    template_name = "emissions/scope1_list.html"
    filter_fields = ["year", "status"]
    search_fields = []
    default_sort = "-year"

    def get_queryset(self):
        self.company = get_object_or_404(Companies, pk=self.kwargs.get("company_id"))
        qs = self.model.objects.filter(company=self.company).select_related("company")

        self.current_sort = self.request.GET.get("sort", self.default_sort)
        self.current_year = self.request.GET.get("year", "")
        self.current_status = self.request.GET.get("status", "")
        self.current_search = self.request.GET.get("q", "")

        if self.current_year:
            qs = qs.filter(year=self.current_year)
        if self.current_status:
            qs = qs.filter(status=self.current_status)
        if self.current_search and self.search_fields:

            search_query = Q()
            for field in self.search_fields:
                search_query |= Q(**{f"{field}__icontains": self.current_search})
            qs = qs.filter(search_query)

        return qs.order_by(self.current_sort)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        status_choices = (
            getattr(self.model._meta.get_field("status"), "choices", [])
            if hasattr(self.model, "status")
            else []
        )
        context.update(
            {
                "company": self.company,
                "model_verbose_name": self.model._meta.verbose_name,
                "model_verbose_name_plural": self.model._meta.verbose_name_plural,
                "current_sort": self.current_sort,
                "current_year": self.current_year,
                "current_status": self.current_status,
                "current_search": self.current_search,
                "status_choices": status_choices,
                "add_url_name": f"emissions:{self.model._meta.model_name}-add",
                "edit_url_name": f"emissions:{self.model._meta.model_name}-edit",
                "delete_url_name": f"emissions:{self.model._meta.model_name}-delete",
            }
        )
        query_params = self.request.GET.copy()
        if "page" in query_params:
            del query_params["page"]

        clean_query_string = query_params.urlencode()
        context["query_string"] = f"{clean_query_string}&" if clean_query_string else ""

        return context

    def get(self, request, *args, **kwargs):
        if "export" in request.GET:
            return self.export_to_excel()
        return super().get(request, *args, **kwargs)

    def export_to_excel(self):
        """Eksportuje wyfiltrowany QuerySet do pliku .xlsx."""
        qs = self.get_queryset()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(self.model._meta.verbose_name_plural)[:31]
        fields = [f for f in self.model._meta.fields if f.name not in ["id", "company"]]
        ws.append([f.verbose_name.title() for f in fields])

        for obj in qs:
            row = []
            for field in fields:
                value = getattr(obj, field.name)
                if hasattr(obj, f"get_{field.name}_display"):
                    value = getattr(obj, f"get_{field.name}_display")()
                row.append(str(value) if value is not None else "")
            ws.append(row)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"export_{self.model._meta.model_name}_{self.company.pk}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class Scope1DeleteMixin(LoginRequiredMixin):
    """Wspólna logika dla usuwania rekordów z Zakresu 1.
    Dynamicznie określa adres przekierowania powrotnego na listę.
    """

    def get_success_url(self):
        return reverse_lazy(
            f"emissions:{self.model._meta.model_name}-list",
            kwargs={"company_id": self.kwargs.get("company_id")},
        )

    def delete(self, request, *args, **kwargs):
        messages.success(
            self.request,
            f"Pomyślnie usunięto wpis z: {self.model._meta.verbose_name}",
        )
        return super().delete(request, *args, **kwargs)


class StationaryCombustionDeleteView(Scope1DeleteMixin, DeleteView):
    """Widok usuwania rekordów dla spalania stacjonarnego."""

    model = StationaryCombustion


class StationaryCombustionListView(EmissionListMixin):
    model = StationaryCombustion
    search_fields = ["fuel", "installation"]


class StationaryCombustionCreateView(Scope1CreateMixin, CreateView):
    model = StationaryCombustion
    form_class = StationaryCombustionForm


class StationaryCombustionUpdateView(Scope1CreateMixin, UpdateView):
    model = StationaryCombustion
    form_class = StationaryCombustionForm


class MobileCombustionCreateView(Scope1CreateMixin, CreateView):
    model = MobileCombustion
    form_class = MobileCombustionForm


class MobileCombustionListView(EmissionListMixin):
    model = MobileCombustion
    search_fields = ["fuel", "vehicle"]


class MobileCombustionUpdateView(Scope1CreateMixin, UpdateView):
    model = MobileCombustion
    form_class = MobileCombustionForm


class MobileCombustionDeleteView(Scope1DeleteMixin, DeleteView):
    model = MobileCombustion


class ProcessEmissionCreateView(Scope1CreateMixin, CreateView):
    model = ProcessEmission
    form_class = ProcessEmissionForm


class ProcessEmissionListView(EmissionListMixin):
    model = ProcessEmission
    search_fields = ["process", "product"]


class ProcessEmissionUpdateView(Scope1CreateMixin, UpdateView):
    model = ProcessEmission
    form_class = ProcessEmissionForm


class ProcessEmissionDeleteView(Scope1DeleteMixin, DeleteView):
    model = ProcessEmission


class FugitiveEmissionCreateView(Scope1CreateMixin, CreateView):
    model = FugitiveEmission
    form_class = FugitiveEmissionForm


class FugitiveEmissionListView(EmissionListMixin):
    model = FugitiveEmission
    search_fields = ["installation", "product"]


class FugitiveEmissionUpdateView(Scope1CreateMixin, UpdateView):
    model = FugitiveEmission
    form_class = FugitiveEmissionForm


class FugitiveEmissionDeleteView(Scope1DeleteMixin, DeleteView):
    model = FugitiveEmission


class DashboardView(LoginRequiredMixin, TemplateView):
    """
    Główny widok nawigacyjny kalkulatora emisji (Dashboard).
    Służy jako menu wyboru zakresów i kategorii.
    """

    template_name = "emissions/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company_id = self.kwargs.get("company_id")
        company = get_object_or_404(Companies, pk=company_id)
        context["company"] = company

        return context


class EmissionFactorListView(LoginRequiredMixin, ListView):
    model = EmissionFactor
    template_name = "emissions/factor_list.html"
    paginate_by = 15

    def get_missing_factors(self):
        """Skanuje całą bazę i zwraca zbiór (rok, nazwa_paliwa), dla których brakuje wskaźnika."""
        required = set()

        for model in [StationaryCombustion, MobileCombustion]:
            for year, factor_name in model.objects.values_list(
                "year", "fuel"
            ).distinct():
                if year and factor_name:
                    required.add((year, factor_name))

        for year, factor_name in ProcessEmission.objects.values_list(
            "year", "process"
        ).distinct():
            if year and factor_name:
                required.add((year, factor_name))

        for year, factor_name in FugitiveEmission.objects.values_list(
            "year", "product"
        ).distinct():
            if year and factor_name:
                required.add((year, factor_name))

        for model in [EnergyConsumption, EnergyPurchased, EnergyProduced, EnergySold]:
            for year, factor_name in model.objects.values_list(
                "year", "energy_type"
            ).distinct():
                if year and factor_name:
                    required.add((year, factor_name))

        existing = set(EmissionFactor.objects.values_list("year", "factor_name"))

        return required - existing

    def get_queryset(self):
        qs = super().get_queryset().order_by("-year", "factor_name")

        self.current_year = self.request.GET.get("year", "")
        self.current_search = self.request.GET.get("q", "")
        self.show_missing = self.request.GET.get("status") == "missing"

        if self.show_missing and (
            self.request.user.is_superuser
            or getattr(self.request.user, "role", "") == "admin"
        ):
            missing_set = self.get_missing_factors()
            virtual_factors = []

            for year, name in sorted(list(missing_set), key=lambda x: (-x[0], x[1])):
                if self.current_year and str(year) != self.current_year:
                    continue
                if (
                    self.current_search
                    and self.current_search.lower() not in name.lower()
                ):
                    continue

                dummy = EmissionFactor(year=year, factor_name=name)
                dummy.is_missing = True
                virtual_factors.append(dummy)

            return virtual_factors

        if self.current_year:
            qs = qs.filter(year=self.current_year)
        if self.current_search:
            qs = qs.filter(factor_name__icontains=self.current_search)

        for obj in qs:
            obj.is_missing = False

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["current_year"] = self.current_year
        context["current_search"] = self.current_search
        context["show_missing"] = self.show_missing
        context["is_admin"] = (
            self.request.user.is_superuser
            or getattr(self.request.user, "role", "") == "admin"
        )

        query_params = self.request.GET.copy()
        if "page" in query_params:
            del query_params["page"]
        clean_qs = query_params.urlencode()
        context["query_string"] = f"{clean_qs}&" if clean_qs else ""

        return context


class EmissionFactorCreateView(LoginRequiredMixin, CreateView):
    model = EmissionFactor
    form_class = EmissionFactorForm
    template_name = "emissions/factor_form.html"
    success_url = reverse_lazy("emissions:factor-list")

    def get_initial(self):
        initial = super().get_initial()
        if "year" in self.request.GET:
            initial["year"] = self.request.GET.get("year")
        if "factor_name" in self.request.GET:
            initial["factor_name"] = self.request.GET.get("factor_name")

        # Domyślnie podpowiadamy Polskę, bo tam działamy najczęściej
        initial["country"] = "PL"
        return initial

    def form_valid(self, form):
        messages.success(
            self.request,
            "Pomyślnie dodano nowy wskaźnik emisji. Możesz teraz przeliczyć brakujące rekordy!",
        )
        return super().form_valid(form)


class EmissionFactorUpdateView(LoginRequiredMixin, UpdateView):
    model = EmissionFactor
    form_class = EmissionFactorForm
    template_name = "emissions/factor_form.html"
    success_url = reverse_lazy("emissions:factor-list")

    def form_valid(self, form):
        messages.success(self.request, "Zaktualizowano wskaźnik emisji.")
        return super().form_valid(form)


class EmissionFactorDeleteView(LoginRequiredMixin, DeleteView):
    model = EmissionFactor
    template_name = "emissions/factor_confirm_delete.html"
    success_url = reverse_lazy("emissions:factor-list")

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Usunięto wskaźnik emisji z bazy.")
        return super().delete(request, *args, **kwargs)


class EnergyConsumptionTemplateDownloadView(View):
    """Pobiera szablon XLSX do wgrania danych."""

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Zużycie energii"

        headers = ['year', 'company', 'energy_source', 'energy_type', 'amount', 'unit', 'source']
        ws.append(headers)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="szablon_energia.xlsx"'
        wb.save(response)
        return response


class EnergyConsumptionImportView(FormView):
    """Widok importu danych z pliku XLSX."""

    template_name = 'emissions/energy_consumption_import.html'
    form_class = EnergyConsumptionImportForm  # tworzymy ten formularz
    success_url = reverse_lazy('energy_consumption_list')

    def form_valid(self, form):
        file = form.cleaned_data['file']

        # Walidacja pliku
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            messages.error(self.request, f'Błąd wczytywania pliku: {e}')
            return self.form_invalid(form)

        # Pobierz nagłówki
        headers = [cell.value for cell in ws[1]]
        expected_headers = ['year', 'company', 'energy_source', 'energy_type', 'amount', 'unit', 'source']

        if headers != expected_headers:
            messages.error(self.request, f'Niepoprawna struktura pliku. Oczekiwane kolumny: {expected_headers}')
            return self.form_invalid(form)

        # Policz rekordy
        records_count = ws.max_row - 1
        context = self.get_context_data(form=form)
        context['records_count'] = records_count
        context['confirm'] = True

        # Jeśli to potwierdzenie — zapisz dane
        if 'confirm' in self.request.POST:
            imported = 0
            duplicates = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                year, company, energy_source, energy_type, amount, unit, source = row

                # Sprawdź duplikat
                exists = EnergyConsumption.objects.filter(
                    year=year,
                    company=company,
                    energy_source=energy_source,
                    energy_type=energy_type
                ).exists()

                if exists:
                    duplicates += 1
                    continue

                # Dodaj do bazy
                EnergyConsumption.objects.create(
                    year=year,
                    company=company,
                    energy_source=energy_source,
                    energy_type=energy_type,
                    amount=amount,
                    unit=unit,
                    source=source
                )
                imported += 1

            messages.success(
                self.request,
                f'Zaimportowano {imported} rekordów. Pominięto {duplicates} duplikatów.'
            )
            return redirect(self.success_url)

        return self.render_to_response(context)


# ===== ENERGY PURCHASED IMPORT =====

class EnergyPurchasedTemplateDownloadView(View):
    """Pobiera szablon XLSX dla zakupionej energii."""

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Zakupiona energia"

        headers = ['year', 'company', 'energy_type', 'amount', 'unit', 'trader', 'source']
        ws.append(headers)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="szablon_energia_zakupiona.xlsx"'
        wb.save(response)
        return response


class EnergyPurchasedImportView(FormView):
    """Widok importu danych zakupionej energii."""

    template_name = 'emissions/energy_purchased_import.html'
    form_class = EnergyPurchasedImportForm
    success_url = reverse_lazy('energy_purchased_list')

    def form_valid(self, form):
        file = form.cleaned_data['file']

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            messages.error(self.request, f'Błąd wczytywania pliku: {e}')
            return self.form_invalid(form)

        headers = [cell.value for cell in ws[1]]
        expected_headers = ['year', 'company', 'energy_type', 'amount', 'unit', 'trader', 'source']

        if headers != expected_headers:
            messages.error(self.request, f'Niepoprawna struktura pliku. Oczekiwane kolumny: {expected_headers}')
            return self.form_invalid(form)

        records_count = ws.max_row - 1
        context = self.get_context_data(form=form)
        context['records_count'] = records_count
        context['confirm'] = True

        if 'confirm' in self.request.POST:
            imported = 0
            duplicates = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                year, company, energy_type, amount, unit, trader, source = row

                exists = EnergyPurchased.objects.filter(
                    year=year,
                    company=company,
                    energy_type=energy_type
                ).exists()

                if exists:
                    duplicates += 1
                    continue

                EnergyPurchased.objects.create(
                    year=year,
                    company=company,
                    energy_type=energy_type,
                    amount=amount,
                    unit=unit,
                    trader=trader,
                    source=source
                )
                imported += 1

            messages.success(
                self.request,
                f'Zaimportowano {imported} rekordów. Pominięto {duplicates} duplikatów.'
            )
            return redirect(self.success_url)

        return self.render_to_response(context)


# ===== ENERGY PRODUCED IMPORT =====

class EnergyProducedTemplateDownloadView(View):
    """Pobiera szablon XLSX dla wyprodukowanej energii."""

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Wyprodukowana energia"

        headers = ['year', 'company', 'energy_type', 'amount', 'unit', 'installation', 'source']
        ws.append(headers)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="szablon_energia_wyprodukowana.xlsx"'
        wb.save(response)
        return response


class EnergyProducedImportView(FormView):
    """Widok importu danych wyprodukowanej energii."""

    template_name = 'emissions/energy_produced_import.html'
    form_class = EnergyProducedImportForm
    success_url = reverse_lazy('energy_produced_list')

    def form_valid(self, form):
        file = form.cleaned_data['file']

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            messages.error(self.request, f'Błąd wczytywania pliku: {e}')
            return self.form_invalid(form)

        headers = [cell.value for cell in ws[1]]
        expected_headers = ['year', 'company', 'energy_type', 'amount', 'unit', 'installation', 'source']

        if headers != expected_headers:
            messages.error(self.request, f'Niepoprawna struktura pliku. Oczekiwane kolumny: {expected_headers}')
            return self.form_invalid(form)

        records_count = ws.max_row - 1
        context = self.get_context_data(form=form)
        context['records_count'] = records_count
        context['confirm'] = True

        if 'confirm' in self.request.POST:
            imported = 0
            duplicates = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                year, company, energy_type, amount, unit, installation, source = row

                exists = EnergyProduced.objects.filter(
                    year=year,
                    company=company,
                    energy_type=energy_type
                ).exists()

                if exists:
                    duplicates += 1
                    continue

                EnergyProduced.objects.create(
                    year=year,
                    company=company,
                    energy_type=energy_type,
                    amount=amount,
                    unit=unit,
                    installation=installation,
                    source=source
                )
                imported += 1

            messages.success(
                self.request,
                f'Zaimportowano {imported} rekordów. Pominięto {duplicates} duplikatów.'
            )
            return redirect(self.success_url)

        return self.render_to_response(context)


# ===== ENERGY SOLD IMPORT =====

class EnergySoldTemplateDownloadView(View):
    """Pobiera szablon XLSX dla sprzedanej energii."""

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sprzedana energia"

        headers = ['year', 'company', 'energy_type', 'amount', 'unit', 'customer', 'source']
        ws.append(headers)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="szablon_energia_sprzedana.xlsx"'
        wb.save(response)
        return response


class EnergySoldImportView(FormView):
    """Widok importu danych sprzedanej energii."""

    template_name = 'emissions/energy_sold_import.html'
    form_class = EnergySoldImportForm
    success_url = reverse_lazy('energy_sold_list')

    def form_valid(self, form):
        file = form.cleaned_data['file']

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            messages.error(self.request, f'Błąd wczytywania pliku: {e}')
            return self.form_invalid(form)

        headers = [cell.value for cell in ws[1]]
        expected_headers = ['year', 'company', 'energy_type', 'amount', 'unit', 'customer', 'source']

        if headers != expected_headers:
            messages.error(self.request, f'Niepoprawna struktura pliku. Oczekiwane kolumny: {expected_headers}')
            return self.form_invalid(form)

        records_count = ws.max_row - 1
        context = self.get_context_data(form=form)
        context['records_count'] = records_count
        context['confirm'] = True

        if 'confirm' in self.request.POST:
            imported = 0
            duplicates = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                year, company, energy_type, amount, unit, customer, source = row

                exists = EnergySold.objects.filter(
                    year=year,
                    company=company,
                    energy_type=energy_type
                ).exists()

                if exists:
                    duplicates += 1
                    continue

                EnergySold.objects.create(
                    year=year,
                    company=company,
                    energy_type=energy_type,
                    amount=amount,
                    unit=unit,
                    customer=customer,
                    source=source
                )
                imported += 1

            messages.success(
                self.request,
                f'Zaimportowano {imported} rekordów. Pominięto {duplicates} duplikatów.'
            )
            return redirect(self.success_url)

        return self.render_to_response(context)
