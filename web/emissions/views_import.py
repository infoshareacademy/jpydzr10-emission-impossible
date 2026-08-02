"""
Widoki importu danych dla modeli Scope 1.

Oddzielony od views.py dla czytelności. Importowany bezpośrednio w urls.py.

Architektura:
  Scope1TemplateDownloadMixin  — pobieranie szablonu XLSX z nagłówkami i przykładem
  Scope1ImportMixin            — dwukrokowy import: upload → podgląd → potwierdzenie

Konkretne widoki (po 2 na model):
  Stationary:  StationaryCombustionTemplateDownloadView, StationaryCombustionImportView
  Mobile:      MobileCombustionTemplateDownloadView,     MobileCombustionImportView
  Process:     ProcessEmissionTemplateDownloadView,       ProcessEmissionImportView
  Fugitive:    FugitiveEmissionTemplateDownloadView,      FugitiveEmissionImportView
"""

from __future__ import annotations

import logging

import openpyxl
from companies.models import Companies
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views import View
from openpyxl.styles import Alignment, Font, PatternFill

from .importers import (
    FugitiveEmissionImporter,
    MobileCombustionImporter,
    ProcessEmissionImporter,
    StationaryCombustionImporter,
)
from .models import (
    FugitiveEmission,
    MobileCombustion,
    ProcessEmission,
    StationaryCombustion,
)

logger = logging.getLogger(__name__)

# Limity bezpieczeństwa pliku
MAX_FILE_SIZE_MB = 5
ALLOWED_EXTENSION = ".xlsx"
ALLOWED_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Kolory szablonu XLSX
COLOR_HEADER_BG = "1a4a52"  # ciemny zielony
COLOR_HEADER_FG = "FFFFFF"  # biały
COLOR_EXAMPLE_BG = "EEF8F0"  # jasny zielony
COLOR_INFO_BG = "FFFBEB"  # jasny żółty
COLOR_INFO_FG = "7C6A00"  # ciemny żółty


# ===========================================================================
# MIXIN: Pobieranie szablonu XLSX
# ===========================================================================


class Scope1TemplateDownloadMixin(LoginRequiredMixin, View):
    """
    Generuje i wysyła szablon XLSX do wypełnienia danymi.
    Każdy konkretny widok definiuje:
      importer_class  — klasa importera (dostarcza nagłówki, przykład, instrukcje)
      sheet_name      — nazwa arkusza
      filename        — nazwa pliku do pobrania
    """

    importer_class = None
    sheet_name: str = "Import"
    filename: str = "szablon.xlsx"

    def get(self, request, company_id: int, **kwargs) -> HttpResponse:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.sheet_name[:31]  # Excel max 31 znaków

        headers = self.importer_class.EXPECTED_HEADERS
        example = self.importer_class.EXAMPLE_ROW
        instructions = self.importer_class.INSTRUCTIONS
        allowed_units = self.importer_class.ALLOWED_UNITS

        # ── Styl wiersza nagłówkowego ────────────────────────────────────────
        header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        header_font = Font(bold=True, color=COLOR_HEADER_FG, size=11)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(headers)
        ws.row_dimensions[1].height = 22
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # ── Wiersz przykładowy ───────────────────────────────────────────────
        ws.append(example)
        example_fill = PatternFill("solid", fgColor=COLOR_EXAMPLE_BG)
        example_font = Font(italic=True, size=10, color="2D6A4F")
        for cell in ws[2]:
            cell.fill = example_fill
            cell.font = example_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # ── Wiersz z instrukcjami ────────────────────────────────────────────
        ws.append([])  # pusty separator
        info_row_idx = 4
        info_font = Font(size=9, color=COLOR_INFO_FG)
        info_fill = PatternFill("solid", fgColor=COLOR_INFO_BG)

        for i, instruction in enumerate(instructions):
            row_idx = info_row_idx + i
            ws.cell(row=row_idx, column=1, value=f"ℹ  {instruction}")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = info_fill
                cell.font = info_font
                cell.alignment = Alignment(wrap_text=True)

        # ── Automatyczna szerokość kolumn ────────────────────────────────────
        col_widths = {
            "rok": 8,
            "paliwo": 28,
            "pojazd": 32,
            "instalacja": 32,
            "proces": 28,
            "produkt": 28,
            "czynnik": 20,
            "ilosc": 14,
            "jednostka": 14,
            "zrodlo": 30,
        }
        for idx, header in enumerate(headers, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(idx)
            ].width = col_widths.get(header, 20)

        # ── Zamrożenie nagłówka ──────────────────────────────────────────────
        ws.freeze_panes = "A2"

        # ── Odpowiedź HTTP ───────────────────────────────────────────────────
        response = HttpResponse(
            content_type="application/"
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{self.filename}"'
        wb.save(response)
        logger.info(
            "Szablon XLSX '%s' pobrany przez %s (firma_id=%s)",
            self.filename,
            request.user,
            company_id,
        )
        return response


# ===========================================================================
# MIXIN: Dwukrokowy import XLSX
# ===========================================================================


class Scope1ImportMixin(LoginRequiredMixin, View):
    """
    Dwukrokowy import danych Scope 1 z pliku XLSX.

    Krok 1 (GET / POST action='upload'):
      - Wyświetla formularz wgrania pliku
      - Po wgraniu: parsuje, waliduje, zapisuje prawidłowe wiersze w sesji
      - Renderuje tabelę podglądu ze statusem każdego wiersza

    Krok 2 (POST action='confirm'):
      - Odczytuje dane z sesji
      - Zapisuje atomowo wszystkie prawidłowe wiersze
      - Przekierowuje na listę z komunikatem

    POST action='cancel':
      - Czyści dane sesji
      - Przekierowuje na listę
    """

    template_name = "emissions/scope1_import.html"
    importer_class = None
    model = None

    # ── Pomocnicze ───────────────────────────────────────────────────────────

    def _session_key(self, company_id: int) -> str:
        return f"scope1_import__{self.model._meta.model_name}__{company_id}"

    def _get_company(self) -> Companies:
        return get_object_or_404(Companies, pk=self.kwargs["company_id"])

    def _model_name(self) -> str:
        return self.model._meta.model_name

    def _list_url(self, company_id: int) -> str:
        return reverse(
            f"emissions:{self._model_name()}-list",
            kwargs={"company_id": company_id},
        )

    def _import_url(self, company_id: int) -> str:
        return reverse(
            f"emissions:{self._model_name()}-import",
            kwargs={"company_id": company_id},
        )

    def _template_url(self, company_id: int) -> str:
        return reverse(
            f"emissions:{self._model_name()}-template",
            kwargs={"company_id": company_id},
        )

    def _base_ctx(self, company: Companies, **extra) -> dict:
        return {
            "company": company,
            "model_verbose_name": self.model._meta.verbose_name,
            "model_verbose_name_plural": self.model._meta.verbose_name_plural,
            "importer_headers": self.importer_class.EXPECTED_HEADERS,
            "allowed_units": self.importer_class.ALLOWED_UNITS,
            "instructions": self.importer_class.INSTRUCTIONS,
            "max_rows": self.importer_class.MAX_ROWS,
            "max_size_mb": MAX_FILE_SIZE_MB,
            "list_url": self._list_url(company.pk),
            "import_url": self._import_url(company.pk),
            "template_url": self._template_url(company.pk),
            **extra,
        }

    # ── GET — wyświetl formularz wgrania ─────────────────────────────────────

    def get(self, request, company_id: int, **kwargs):
        company = self._get_company()
        # Czyścimy ewentualne stare dane sesji (idempotentne)
        request.session.pop(self._session_key(company.pk), None)
        return render(
            request,
            self.template_name,
            self._base_ctx(company, step="upload"),
        )

    # ── POST — dispatch do odpowiedniego kroku ───────────────────────────────

    def post(self, request, company_id: int, **kwargs):
        company = self._get_company()
        action = request.POST.get("action", "upload")

        if action == "confirm":
            return self._handle_confirm(request, company)
        elif action == "cancel":
            return self._handle_cancel(request, company)
        else:
            return self._handle_upload(request, company)

    # ── Krok 1: Upload i walidacja ───────────────────────────────────────────

    def _handle_upload(self, request, company: Companies):
        file = request.FILES.get("file")

        # Brak pliku
        if not file:
            messages.error(
                request, "Nie wybrano pliku. Wybierz plik XLSX i spróbuj ponownie."
            )
            return render(
                request,
                self.template_name,
                self._base_ctx(company, step="upload"),
            )

        # Walidacja rozszerzenia
        if not file.name.lower().endswith(ALLOWED_EXTENSION):
            messages.error(
                request,
                f"Niedozwolony format pliku: '{file.name}'. "
                f"Wymagany format: {ALLOWED_EXTENSION.upper()}.",
            )
            return render(
                request,
                self.template_name,
                self._base_ctx(company, step="upload"),
            )

        # Walidacja rozmiaru
        if file.size > MAX_FILE_SIZE_MB * 1024 * 1024:
            messages.error(
                request,
                f"Plik jest za duży ({file.size / 1024 / 1024:.1f} MB). "
                f"Maksymalny dozwolony rozmiar: {MAX_FILE_SIZE_MB} MB.",
            )
            return render(
                request,
                self.template_name,
                self._base_ctx(company, step="upload"),
            )

        # Parsowanie i walidacja
        importer = self.importer_class(company=company, user=request.user)
        ok = importer.parse(file)

        if not ok:
            # Błędy strukturalne (złe nagłówki, nieczytelny plik itp.)
            logger.warning(
                "Import XLSX nieudany (błędy struktury): model=%s user=%s errors=%s",
                self._model_name(),
                request.user,
                importer.parse_errors,
            )
            return render(
                request,
                self.template_name,
                self._base_ctx(
                    company,
                    step="upload",
                    parse_errors=importer.parse_errors,
                ),
            )

        # Przygotowanie danych podglądu (JSON-safe dict dla szablonu)
        preview_rows = self._build_preview_rows(importer)

        # Zapis prawidłowych wierszy w sesji (tylko jeśli są)
        valid_count = len(importer.valid_rows)
        if valid_count > 0:
            request.session[self._session_key(company.pk)] = (
                importer.to_session_payload()
            )
            # Sesja: wymuszamy zapis (niektóre backendy są leniwe)
            request.session.modified = True

        logger.info(
            "Import XLSX podgląd: model=%s user=%s firma=%s "
            "łącznie=%d prawidłowych=%d błędnych=%d",
            self._model_name(),
            request.user,
            company.pk,
            len(importer.rows),
            valid_count,
            len(importer.invalid_rows),
        )

        return render(
            request,
            self.template_name,
            self._base_ctx(
                company,
                step="preview",
                preview_rows=preview_rows,
                headers=self.importer_class.EXPECTED_HEADERS,
                total_count=len(importer.rows),
                valid_count=valid_count,
                invalid_count=len(importer.invalid_rows),
            ),
        )

    # ── Krok 2: Potwierdzenie i zapis ────────────────────────────────────────

    def _handle_confirm(self, request, company: Companies):
        session_key = self._session_key(company.pk)
        payload = request.session.get(session_key)

        if not payload:
            messages.warning(
                request,
                "Dane importu wygasły lub sesja została wyczyszczona. "
                "Proszę wgrać plik ponownie.",
            )
            return redirect(self._import_url(company.pk))

        importer = self.importer_class(company=company, user=request.user)
        try:
            saved = importer.save_from_payload(payload)
        except Exception as exc:
            logger.error(
                "Import XLSX BŁĄD ZAPISU: model=%s user=%s firma=%s error=%s",
                self._model_name(),
                request.user,
                company.pk,
                exc,
            )
            messages.error(
                request,
                f"Błąd podczas zapisu danych — żaden rekord nie został zapisany "
                f"(transakcja wycofana). Szczegóły: {exc}",
            )
            return redirect(self._import_url(company.pk))

        # Sukces — czyścimy sesję
        request.session.pop(session_key, None)

        logger.info(
            "Import XLSX SUKCES: model=%s user=%s firma=%s zapisano=%d",
            self._model_name(),
            request.user,
            company.pk,
            saved,
        )

        messages.success(
            request,
            f"✅ Pomyślnie zaimportowano {saved} rekordów "
            f"({self.model._meta.verbose_name_plural}). "
            f"Wszystkie rekordy mają status Roboczy i oczekują na weryfikację.",
        )
        return redirect(self._list_url(company.pk))

    # ── Anulowanie ───────────────────────────────────────────────────────────

    def _handle_cancel(self, request, company: Companies):
        request.session.pop(self._session_key(company.pk), None)
        messages.info(request, "Import anulowany.")
        return redirect(self._list_url(company.pk))

    # ── Budowanie danych podglądu dla szablonu ───────────────────────────────

    @staticmethod
    def _build_preview_rows(importer) -> list[dict]:
        """
        Przekształca ImportRow objects w proste słowniki przyjazne szablonowi.
        values = lista wartości raw w kolejności kolumn (do iteracji w templatce).
        """
        headers = importer.EXPECTED_HEADERS
        result = []
        for row in importer.rows:
            result.append(
                {
                    "row_num": row.row_num,
                    "values": [
                        str(row.raw.get(h, "")) if row.raw.get(h) is not None else ""
                        for h in headers
                    ],
                    "is_valid": row.is_valid,
                    "errors": row.errors,
                }
            )
        return result


# ===========================================================================
# KONKRETNE WIDOKI — Spalanie stacjonarne
# ===========================================================================


class StationaryCombustionTemplateDownloadView(Scope1TemplateDownloadMixin):
    importer_class = StationaryCombustionImporter
    sheet_name = "Spalanie stacjonarne"
    filename = "szablon_spalanie_stacjonarne.xlsx"


class StationaryCombustionImportView(Scope1ImportMixin):
    model = StationaryCombustion
    importer_class = StationaryCombustionImporter


# ===========================================================================
# KONKRETNE WIDOKI — Spalanie mobilne
# ===========================================================================


class MobileCombustionTemplateDownloadView(Scope1TemplateDownloadMixin):
    importer_class = MobileCombustionImporter
    sheet_name = "Spalanie mobilne"
    filename = "szablon_spalanie_mobilne.xlsx"


class MobileCombustionImportView(Scope1ImportMixin):
    model = MobileCombustion
    importer_class = MobileCombustionImporter


# ===========================================================================
# KONKRETNE WIDOKI — Emisje procesowe
# ===========================================================================


class ProcessEmissionTemplateDownloadView(Scope1TemplateDownloadMixin):
    importer_class = ProcessEmissionImporter
    sheet_name = "Emisje procesowe"
    filename = "szablon_emisje_procesowe.xlsx"


class ProcessEmissionImportView(Scope1ImportMixin):
    model = ProcessEmission
    importer_class = ProcessEmissionImporter


# ===========================================================================
# KONKRETNE WIDOKI — Emisje niezorganizowane
# ===========================================================================


class FugitiveEmissionTemplateDownloadView(Scope1TemplateDownloadMixin):
    importer_class = FugitiveEmissionImporter
    sheet_name = "Emisje niezorganizowane"
    filename = "szablon_emisje_niezorganizowane.xlsx"


class FugitiveEmissionImportView(Scope1ImportMixin):
    model = FugitiveEmission
    importer_class = FugitiveEmissionImporter
